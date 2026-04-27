from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import uuid
import base64
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal
from datetime import datetime, timezone
import openpyxl
from openpyxl import Workbook
import qrcode


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Buldhana Police Bandobast API")
api_router = APIRouter(prefix="/api")


# ========================= MODELS =========================

StaffType = Literal["officer", "amaldar", "home_guard"]

OFFICER_RANKS = ["ASP", "Dy.SP", "PI", "API", "PSI"]
AMALDAR_RANKS = ["ASI", "HC", "NPC", "PC", "LPC"]
HOME_GUARD_RANKS = ["Home Guard"]


class Staff(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    staff_type: StaffType
    rank: str
    bakkal_no: str = ""  # optional for officers
    name: str
    posting: str = ""
    mobile: str = ""
    photo: str = ""  # base64 data url or URL
    gender: Literal["Male", "Female", "Other"] = "Male"
    district: str = "Buldhana"
    category: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class StaffCreate(BaseModel):
    staff_type: StaffType
    rank: str
    bakkal_no: Optional[str] = ""
    name: str
    posting: Optional[str] = ""
    mobile: Optional[str] = ""
    photo: Optional[str] = ""
    gender: Optional[str] = "Male"
    district: Optional[str] = "Buldhana"
    category: Optional[str] = ""


class StaffUpdate(BaseModel):
    rank: Optional[str] = None
    name: Optional[str] = None
    posting: Optional[str] = None
    mobile: Optional[str] = None
    photo: Optional[str] = None
    gender: Optional[str] = None
    district: Optional[str] = None
    category: Optional[str] = None


class BandobastPoint(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    point_name: str
    seq: int = 0
    req_officer: int = 0
    req_amaldar: int = 0
    req_female_amaldar: int = 0
    req_home_guard: int = 0
    equipment: List[str] = []
    sector: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    suchana: str = ""
    is_reserved: bool = False


class Bandobast(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    year: int
    date: str  # ISO
    name: str
    spot: str = ""
    ps_name: str = ""
    in_charge: str = ""
    has_other_district: bool = False
    other_district_staff: List[dict] = []
    points: List[BandobastPoint] = []
    selected_staff_ids: List[str] = []
    allotments: dict = {}  # point_id -> [staff_id,...]
    equipment_assignments: dict = {}  # point_id -> {staff_id: equipment_name}
    status: Literal["draft", "deployed"] = "draft"
    deleted: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class BandobastCreate(BaseModel):
    year: int
    date: str
    name: str
    spot: Optional[str] = ""
    ps_name: Optional[str] = ""
    in_charge: Optional[str] = ""
    has_other_district: Optional[bool] = False


class BandobastUpdate(BaseModel):
    year: Optional[int] = None
    date: Optional[str] = None
    name: Optional[str] = None
    spot: Optional[str] = None
    ps_name: Optional[str] = None
    in_charge: Optional[str] = None
    has_other_district: Optional[bool] = None


# ========================= STAFF ENDPOINTS =========================

@api_router.get("/")
async def root():
    return {"message": "Buldhana Police Bandobast API"}


@api_router.get("/staff", response_model=List[Staff])
async def list_staff(staff_type: Optional[StaffType] = None, rank: Optional[str] = None, search: Optional[str] = None):
    query = {}
    if staff_type:
        query["staff_type"] = staff_type
    if rank:
        query["rank"] = rank
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"bakkal_no": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}},
        ]
    items = await db.staff.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return items


@api_router.post("/staff", response_model=Staff)
async def create_staff(payload: StaffCreate):
    # Duplicate check: officers use (name + mobile); others use bakkal_no
    if payload.staff_type == "officer":
        existing = await db.staff.find_one(
            {"staff_type": "officer", "name": payload.name, "mobile": payload.mobile or ""},
            {"_id": 0},
        )
        if existing:
            raise HTTPException(status_code=409, detail="Officer with this Name + Mobile already exists")
    else:
        if not payload.bakkal_no:
            raise HTTPException(status_code=400, detail="Bakkal No is required for this staff type")
        existing = await db.staff.find_one(
            {"bakkal_no": payload.bakkal_no, "staff_type": payload.staff_type},
            {"_id": 0},
        )
        if existing:
            raise HTTPException(status_code=409, detail="Staff with this Bakkal No already exists")
    obj = Staff(**payload.model_dump())
    await db.staff.insert_one(obj.model_dump())
    return obj


@api_router.get("/staff/by-bakkal/{bakkal_no}")
async def get_staff_by_bakkal(bakkal_no: str, staff_type: Optional[StaffType] = None):
    query = {"bakkal_no": bakkal_no}
    if staff_type:
        query["staff_type"] = staff_type
    item = await db.staff.find_one(query, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    return item


@api_router.get("/staff/{staff_id}", response_model=Staff)
async def get_staff(staff_id: str):
    item = await db.staff.find_one({"id": staff_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Staff not found")
    return item


@api_router.patch("/staff/{staff_id}", response_model=Staff)
async def update_staff(staff_id: str, payload: StaffUpdate):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    res = await db.staff.update_one({"id": staff_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    item = await db.staff.find_one({"id": staff_id}, {"_id": 0})
    return item


@api_router.delete("/staff/{staff_id}")
async def delete_staff(staff_id: str):
    res = await db.staff.delete_one({"id": staff_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.delete("/staff/bulk/{staff_type}")
async def delete_all_staff_of_type(staff_type: StaffType):
    """Delete every staff member of the given type (officer/amaldar/home_guard)."""
    res = await db.staff.delete_many({"staff_type": staff_type})
    return {"ok": True, "deleted": res.deleted_count}


@api_router.get("/staff-template/{staff_type}")
async def staff_template(staff_type: StaffType):
    wb = Workbook()
    ws = wb.active
    ws.title = staff_type.upper()
    if staff_type == "officer":
        headers = ["rank", "name", "posting", "mobile", "gender", "district", "category"]
        ws.append(headers)
        ws.append(["PI", "Example Officer", "PS Buldhana", "9999999999", "Male", "Buldhana", "Open"])
    else:
        headers = ["rank", "bakkal_no", "name", "posting", "mobile", "gender", "district", "category"]
        ws.append(headers)
        example_rank = "HC" if staff_type == "amaldar" else "Home Guard"
        ws.append([example_rank, "12345", "Example Name", "PS Buldhana", "9999999999", "Male", "Buldhana", "Open"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{staff_type}_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.post("/staff/import/{staff_type}")
async def import_staff(staff_type: StaffType, file: UploadFile = File(...)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"inserted": 0, "skipped_duplicate": 0, "skipped_missing": 0, "total_rows": 0, "errors": []}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    is_officer = staff_type == "officer"
    required = ["rank", "name"] if is_officer else ["rank", "bakkal_no", "name"]
    for r in required:
        if r not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required column: {r}")
    inserted = 0
    skipped_duplicate = 0
    skipped_missing = 0
    errors = []
    for i, row in enumerate(rows[1:], start=2):
        data = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
        missing = False
        if is_officer:
            if not data.get("name") or not data.get("rank"):
                missing = True
        else:
            if not data.get("bakkal_no") or not data.get("name") or not data.get("rank"):
                missing = True
        if missing:
            if any(data.values()):
                skipped_missing += 1
            continue
        if is_officer:
            existing = await db.staff.find_one(
                {"staff_type": "officer", "name": data.get("name"), "mobile": data.get("mobile", "")},
                {"_id": 0},
            )
        else:
            existing = await db.staff.find_one(
                {"bakkal_no": data["bakkal_no"], "staff_type": staff_type},
                {"_id": 0},
            )
        if existing:
            skipped_duplicate += 1
            continue
        try:
            obj = Staff(
                staff_type=staff_type,
                rank=data.get("rank", ""),
                bakkal_no=data.get("bakkal_no", "") if not is_officer else "",
                name=data.get("name", ""),
                posting=data.get("posting", ""),
                mobile=data.get("mobile", ""),
                gender=data.get("gender", "Male") or "Male",
                district=data.get("district", "Buldhana") or "Buldhana",
                category=data.get("category", ""),
            )
            await db.staff.insert_one(obj.model_dump())
            inserted += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    return {
        "inserted": inserted,
        "skipped_duplicate": skipped_duplicate,
        "skipped_missing": skipped_missing,
        "errors": errors,
        "total_rows": len(rows) - 1,
    }


# ========================= BANDOBAST ENDPOINTS =========================

@api_router.get("/bandobasts", response_model=List[Bandobast])
async def list_bandobasts():
    items = await db.bandobasts.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.get("/bandobasts/deleted", response_model=List[Bandobast])
async def list_deleted_bandobasts():
    items = await db.bandobasts.find({"deleted": True}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.post("/bandobasts", response_model=Bandobast)
async def create_bandobast(payload: BandobastCreate):
    obj = Bandobast(**payload.model_dump())
    await db.bandobasts.insert_one(obj.model_dump())
    return obj


@api_router.get("/bandobasts/{bid}", response_model=Bandobast)
async def get_bandobast(bid: str):
    item = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    return item


@api_router.patch("/bandobasts/{bid}", response_model=Bandobast)
async def update_bandobast(bid: str, payload: BandobastUpdate):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    res = await db.bandobasts.update_one({"id": bid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    item = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    return item


@api_router.delete("/bandobasts/{bid}")
async def delete_bandobast(bid: str):
    """Soft-delete: moves to Deleted Bandobasts tab."""
    res = await db.bandobasts.update_one({"id": bid}, {"$set": {"deleted": True}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True, "soft_deleted": True}


@api_router.post("/bandobasts/{bid}/restore")
async def restore_bandobast(bid: str):
    res = await db.bandobasts.update_one({"id": bid}, {"$set": {"deleted": False}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True, "restored": True}


@api_router.delete("/bandobasts/{bid}/permanent")
async def delete_bandobast_permanent(bid: str):
    res = await db.bandobasts.delete_one({"id": bid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True, "permanent": True}


# ---- Points Template & Import ----
POINT_TEMPLATE_HEADERS = [
    "point_name", "req_officer", "req_amaldar", "req_female_amaldar",
    "req_home_guard", "equipment", "sector", "latitude", "longitude", "suchana",
]


@api_router.get("/bandobast-point-template")
async def bandobast_point_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "POINTS"
    ws.append(POINT_TEMPLATE_HEADERS)
    ws.append([
        "Main Gate", 1, 4, 1, 2,
        "Lathi,Wireless,Barricade",
        "Sector A", 20.5316, 76.1853,
        "Report 30 min prior. Maintain crowd control.",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bandobast_points_template.xlsx"},
    )


@api_router.post("/bandobasts/{bid}/points/import")
async def import_points(bid: str, file: UploadFile = File(...)):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Bandobast not found")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"inserted": 0, "errors": []}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    if "point_name" not in headers:
        raise HTTPException(status_code=400, detail="Missing required column: point_name")
    inserted = 0
    errors = []
    for i, row in enumerate(rows[1:], start=2):
        data = {headers[j]: row[j] for j in range(len(headers)) if j < len(row)}
        name = str(data.get("point_name") or "").strip()
        if not name:
            continue
        def _int(v):
            try:
                return int(v) if v not in (None, "") else 0
            except Exception:
                return 0
        def _float(v):
            try:
                return float(v) if v not in (None, "") else None
            except Exception:
                return None
        equip_raw = data.get("equipment") or ""
        equipment = [e.strip() for e in str(equip_raw).split(",") if e.strip()]
        try:
            pt = BandobastPoint(
                point_name=name,
                req_officer=_int(data.get("req_officer")),
                req_amaldar=_int(data.get("req_amaldar")),
                req_female_amaldar=_int(data.get("req_female_amaldar")),
                req_home_guard=_int(data.get("req_home_guard")),
                equipment=equipment,
                sector=str(data.get("sector") or "").strip(),
                latitude=_float(data.get("latitude")),
                longitude=_float(data.get("longitude")),
                suchana=str(data.get("suchana") or "").strip(),
            )
            await db.bandobasts.update_one({"id": bid}, {"$push": {"points": pt.model_dump()}})
            inserted += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    return {"inserted": inserted, "errors": errors}


@api_router.get("/bandobasts/{bid}/staff/{sid}")
async def get_any_staff(bid: str, sid: str):
    """Resolve a staff ID either from global staff or bandobast's out-district list."""
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Bandobast not found")
    for s in bandobast.get("other_district_staff", []):
        if s.get("id") == sid:
            return s
    item = await db.staff.find_one({"id": sid}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Staff not found")
    return item


# ---- Out of District Staff (scoped to bandobast only) ----
@api_router.post("/bandobasts/{bid}/out-staff/import/{staff_type}")
async def import_out_staff(bid: str, staff_type: StaffType, file: UploadFile = File(...)):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Bandobast not found")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"inserted": 0, "total_rows": 0}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    is_officer = staff_type == "officer"
    required = ["rank", "name"] if is_officer else ["rank", "bakkal_no", "name"]
    for r in required:
        if r not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required column: {r}")
    existing_out = bandobast.get("other_district_staff", [])
    existing_bakkal = {(s.get("bakkal_no"), s.get("staff_type")) for s in existing_out if s.get("bakkal_no")}
    existing_off = {(s.get("name"), s.get("mobile") or "") for s in existing_out if s.get("staff_type") == "officer"}
    inserted = 0
    skipped_missing = 0
    skipped_duplicate = 0
    new_rows = []
    for row in rows[1:]:
        data = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
        if is_officer:
            if not data.get("name") or not data.get("rank"):
                if any(data.values()):
                    skipped_missing += 1
                continue
            key = (data.get("name"), data.get("mobile", ""))
            if key in existing_off:
                skipped_duplicate += 1
                continue
            existing_off.add(key)
        else:
            if not data.get("bakkal_no") or not data.get("name") or not data.get("rank"):
                if any(data.values()):
                    skipped_missing += 1
                continue
            key = (data["bakkal_no"], staff_type)
            if key in existing_bakkal:
                skipped_duplicate += 1
                continue
            existing_bakkal.add(key)
        new_rows.append({
            "id": str(uuid.uuid4()),
            "staff_type": staff_type,
            "rank": data.get("rank", ""),
            "bakkal_no": data.get("bakkal_no", "") if not is_officer else "",
            "name": data.get("name", ""),
            "posting": data.get("posting", ""),
            "mobile": data.get("mobile", ""),
            "gender": data.get("gender", "Male") or "Male",
            "district": data.get("district", "") or "Other",
            "category": data.get("category", ""),
            "is_out_district": True,
        })
        inserted += 1
    if new_rows:
        await db.bandobasts.update_one({"id": bid}, {"$push": {"other_district_staff": {"$each": new_rows}}})
    return {
        "inserted": inserted,
        "skipped_duplicate": skipped_duplicate,
        "skipped_missing": skipped_missing,
        "total_rows": len(rows) - 1,
    }


class OutStaffCreate(BaseModel):
    staff_type: StaffType
    rank: str
    bakkal_no: str
    name: str
    posting: Optional[str] = ""
    mobile: Optional[str] = ""
    gender: Optional[str] = "Male"
    district: Optional[str] = "Other"
    category: Optional[str] = ""


@api_router.post("/bandobasts/{bid}/out-staff")
async def add_out_staff(bid: str, payload: OutStaffCreate):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Bandobast not found")
    existing_list = bandobast.get("other_district_staff", [])
    if payload.staff_type == "officer":
        for s in existing_list:
            if s.get("staff_type") == "officer" and s.get("name") == payload.name and (s.get("mobile") or "") == (payload.mobile or ""):
                raise HTTPException(status_code=409, detail="Officer with this Name + Mobile already exists")
    else:
        if not payload.bakkal_no:
            raise HTTPException(status_code=400, detail="Bakkal No is required for this staff type")
        for s in existing_list:
            if s.get("bakkal_no") == payload.bakkal_no and s.get("staff_type") == payload.staff_type:
                raise HTTPException(status_code=409, detail="Bakkal No already exists for this bandobast")
    obj = {
        "id": str(uuid.uuid4()),
        **payload.model_dump(),
        "bakkal_no": payload.bakkal_no if payload.staff_type != "officer" else "",
        "is_out_district": True,
    }
    await db.bandobasts.update_one({"id": bid}, {"$push": {"other_district_staff": obj}})
    return obj


class OutStaffUpdate(BaseModel):
    rank: Optional[str] = None
    bakkal_no: Optional[str] = None
    name: Optional[str] = None
    posting: Optional[str] = None
    mobile: Optional[str] = None
    gender: Optional[str] = None
    district: Optional[str] = None
    category: Optional[str] = None


@api_router.patch("/bandobasts/{bid}/out-staff/{sid}")
async def update_out_staff(bid: str, sid: str, payload: OutStaffUpdate):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        return {"ok": True}
    set_ops = {f"other_district_staff.$.{k}": v for k, v in update.items()}
    res = await db.bandobasts.update_one(
        {"id": bid, "other_district_staff.id": sid},
        {"$set": set_ops},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.delete("/bandobasts/{bid}/out-staff/{sid}")
async def delete_out_staff(bid: str, sid: str):
    await db.bandobasts.update_one(
        {"id": bid},
        {"$pull": {"other_district_staff": {"id": sid}}},
    )
    # Also remove from selected + allotments
    await db.bandobasts.update_one({"id": bid}, {"$pull": {"selected_staff_ids": sid}})
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if bandobast:
        allot = bandobast.get("allotments", {})
        changed = False
        for pid, sids in list(allot.items()):
            if sid in sids:
                allot[pid] = [x for x in sids if x != sid]
                changed = True
        if changed:
            await db.bandobasts.update_one({"id": bid}, {"$set": {"allotments": allot}})
    return {"ok": True}


# ---- Points ----
@api_router.post("/bandobasts/{bid}/points", response_model=BandobastPoint)
async def add_point(bid: str, point: BandobastPoint):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Not found")
    if not point.id:
        point.id = str(uuid.uuid4())
    if not point.seq:
        existing_seqs = [p.get("seq", 0) for p in bandobast.get("points", []) if not p.get("is_reserved")]
        point.seq = (max(existing_seqs) if existing_seqs else 0) + 1
    await db.bandobasts.update_one({"id": bid}, {"$push": {"points": point.model_dump()}})
    return point


class PointSeqUpdate(BaseModel):
    seq: int


@api_router.patch("/bandobasts/{bid}/points/{pid}/seq")
async def update_point_seq(bid: str, pid: str, payload: PointSeqUpdate):
    res = await db.bandobasts.update_one(
        {"id": bid, "points.id": pid},
        {"$set": {"points.$.seq": payload.seq}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.patch("/bandobasts/{bid}/points/{pid}", response_model=BandobastPoint)
async def update_point(bid: str, pid: str, point: BandobastPoint):
    point.id = pid
    res = await db.bandobasts.update_one(
        {"id": bid, "points.id": pid},
        {"$set": {"points.$": point.model_dump()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return point


@api_router.delete("/bandobasts/{bid}/points/{pid}")
async def delete_point(bid: str, pid: str):
    await db.bandobasts.update_one({"id": bid}, {"$pull": {"points": {"id": pid}}})
    # also remove allotments for this point
    await db.bandobasts.update_one({"id": bid}, {"$unset": {f"allotments.{pid}": ""}})
    return {"ok": True}


# ---- Selected Staff ----
class SelectedStaffPayload(BaseModel):
    staff_ids: List[str]


@api_router.put("/bandobasts/{bid}/selected-staff")
async def set_selected_staff(bid: str, payload: SelectedStaffPayload):
    res = await db.bandobasts.update_one({"id": bid}, {"$set": {"selected_staff_ids": payload.staff_ids}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True, "count": len(payload.staff_ids)}


# ---- Allotment ----
class AllotmentPayload(BaseModel):
    allotments: dict  # point_id -> [staff_id]


@api_router.put("/bandobasts/{bid}/allotments")
async def set_allotments(bid: str, payload: AllotmentPayload):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Not found")
    await db.bandobasts.update_one({"id": bid}, {"$set": {"allotments": payload.allotments}})
    return {"ok": True}


# ---- Deploy ----
@api_router.post("/bandobasts/{bid}/deploy")
async def deploy_bandobast(bid: str):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Not found")
    allot = dict(bandobast.get("allotments", {}))
    selected = set(bandobast.get("selected_staff_ids", []))
    allotted = set()
    for sids in allot.values():
        allotted.update(sids)
    remaining = list(selected - allotted)
    reserved_point_id = None
    for p in bandobast.get("points", []):
        if p.get("is_reserved"):
            reserved_point_id = p["id"]
            break
    if remaining:
        if not reserved_point_id:
            reserved_point_id = str(uuid.uuid4())
            reserved_point = BandobastPoint(
                id=reserved_point_id,
                point_name="Reserved / राखीव",
                is_reserved=True,
                seq=9999,
            ).model_dump()
            await db.bandobasts.update_one({"id": bid}, {"$push": {"points": reserved_point}})
        allot[reserved_point_id] = list(set(allot.get(reserved_point_id, []) + remaining))
        await db.bandobasts.update_one({"id": bid}, {"$set": {"allotments": allot}})
    await db.bandobasts.update_one({"id": bid}, {"$set": {"status": "deployed"}})
    return {"ok": True, "status": "deployed", "reserved_count": len(remaining)}


# ---- Equipment Assignment ----
class EquipmentAssignmentPayload(BaseModel):
    equipment_assignments: dict  # point_id -> {staff_id: equipment_name}


@api_router.put("/bandobasts/{bid}/equipment-assignments")
async def set_equipment_assignments(bid: str, payload: EquipmentAssignmentPayload):
    res = await db.bandobasts.update_one(
        {"id": bid}, {"$set": {"equipment_assignments": payload.equipment_assignments}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


# ---- QR Code ----
@api_router.get("/bandobasts/{bid}/points/{pid}/qr")
async def point_qr(bid: str, pid: str):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Not found")
    point = next((p for p in bandobast.get("points", []) if p["id"] == pid), None)
    if not point:
        raise HTTPException(status_code=404, detail="Point not found")
    lat = point.get("latitude")
    lng = point.get("longitude")

    # Resolve allotted staff names + equipment
    allot = bandobast.get("allotments", {}).get(pid, [])
    eq_map = bandobast.get("equipment_assignments", {}).get(pid, {})
    all_ids = allot
    home = await db.staff.find({"id": {"$in": all_ids}}, {"_id": 0}).to_list(1000) if all_ids else []
    out = [s for s in bandobast.get("other_district_staff", []) if s["id"] in set(all_ids)]
    staff_map = {s["id"]: s for s in home + out}

    officers = []
    amaldars = []
    home_guards = []
    for sid in allot:
        s = staff_map.get(sid)
        if not s:
            continue
        bkl = s.get("bakkal_no") or ""
        eq = eq_map.get(sid)
        line = s.get("name", "")
        extras = [s.get("rank", "")]
        if bkl and s.get("staff_type") != "officer":
            extras.append(f"B{bkl}")
        if s.get("mobile"):
            extras.append(s["mobile"])
        if eq:
            extras.append(f"[{eq}]")
        line = f"{line} ({', '.join([e for e in extras if e])})"
        if s.get("staff_type") == "officer":
            officers.append(line)
        elif s.get("staff_type") == "amaldar":
            amaldars.append(line)
        else:
            home_guards.append(line)

    lines = [
        f"BANDOBAST: {bandobast.get('name', '')}",
        f"DATE: {bandobast.get('date', '')}",
        f"POINT: {point.get('point_name', '')}",
    ]
    if point.get("sector"):
        lines.append(f"SECTOR: {point['sector']}")
    if lat is not None and lng is not None:
        lines.append(f"MAP: https://www.google.com/maps?q={lat},{lng}")
    if point.get("equipment"):
        lines.append(f"EQUIPMENT: {', '.join(point['equipment'])}")
    if officers:
        lines.append("OFFICERS:")
        lines.extend([f"- {o}" for o in officers])
    if amaldars:
        lines.append("AMALDARS:")
        lines.extend([f"- {a}" for a in amaldars])
    if home_guards:
        lines.append("HOME GUARDS:")
        lines.extend([f"- {h}" for h in home_guards])
    if point.get("suchana"):
        lines.append(f"SUCHANA: {point['suchana']}")

    text = "\n".join(lines)
    # Use higher version + lower error correction to accommodate more data
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=6, border=2)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="image/png")


# ---- Goshwara (Summary) ----
@api_router.get("/bandobasts/{bid}/goshwara")
async def goshwara(bid: str):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Not found")
    all_staff_ids = list({sid for sids in bandobast.get("allotments", {}).values() for sid in sids})
    home_staff = await db.staff.find({"id": {"$in": all_staff_ids}}, {"_id": 0}).to_list(5000)
    out_staff = [s for s in bandobast.get("other_district_staff", []) if s["id"] in set(all_staff_ids)]
    staff_map = {s["id"]: s for s in home_staff}
    for s in out_staff:
        staff_map[s["id"]] = s
    point_wise = []
    for p in bandobast.get("points", []):
        assigned = bandobast.get("allotments", {}).get(p["id"], [])
        point_wise.append({
            "point": p,
            "staff": [staff_map[s] for s in assigned if s in staff_map],
        })
    # staff-wise
    staff_wise = []
    for sid, staff in staff_map.items():
        points = []
        for p in bandobast.get("points", []):
            if sid in bandobast.get("allotments", {}).get(p["id"], []):
                points.append(p)
        staff_wise.append({"staff": staff, "points": points})
    return {
        "bandobast": bandobast,
        "point_wise": point_wise,
        "staff_wise": staff_wise,
    }


@api_router.get("/bandobasts/{bid}/export/staff-wise")
async def export_staff_wise(bid: str):
    bandobast = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not bandobast:
        raise HTTPException(status_code=404, detail="Not found")
    all_ids = list({sid for sids in bandobast.get("allotments", {}).values() for sid in sids})
    selected_ids = bandobast.get("selected_staff_ids", [])
    combined = list({*selected_ids, *all_ids})
    home_staff = await db.staff.find({"id": {"$in": combined}}, {"_id": 0}).to_list(5000)
    out_staff = [s for s in bandobast.get("other_district_staff", []) if s["id"] in set(combined)]
    staff_map = {s["id"]: s for s in home_staff}
    for s in out_staff:
        staff_map[s["id"]] = s
    points = bandobast.get("points", [])
    allot = bandobast.get("allotments", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Amaldar-wise"
    ws.append([
        "Sr", "Type", "Rank", "Bakkal No", "Name", "Posting", "Mobile",
        "Gender", "District", "Category", "Allotted Points",
    ])
    i = 1
    for sid in selected_ids or list(staff_map.keys()):
        s = staff_map.get(sid)
        if not s:
            continue
        assigned = [p["point_name"] for p in points if sid in allot.get(p["id"], [])]
        ws.append([
            i,
            s.get("staff_type", ""),
            s.get("rank", ""),
            s.get("bakkal_no", ""),
            s.get("name", ""),
            s.get("posting", ""),
            s.get("mobile", ""),
            s.get("gender", ""),
            s.get("district", ""),
            s.get("category", ""),
            ", ".join(assigned) or "-",
        ])
        i += 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (bandobast.get("name") or "bandobast").replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_amaldar_wise.xlsx"},
    )


# ========================= BANDOBAST ALERT + STAFF APP =========================

class AlertCreate(BaseModel):
    pass  # body unused; bid in path


def _normalize_mobile(m: str) -> str:
    return "".join(ch for ch in (m or "") if ch.isdigit())[-10:]


async def _resolve_staff_for_bandobast(b: dict, sid: str) -> Optional[dict]:
    for s in (b.get("other_district_staff") or []):
        if s.get("id") == sid:
            return s
    home = await db.staff.find_one({"id": sid}, {"_id": 0})
    return home


def _point_for_staff(b: dict, sid: str) -> Optional[dict]:
    for p in (b.get("points") or []):
        if sid in (b.get("allotments", {}).get(p["id"], []) or []):
            return p
    return None


def _co_staff_ids(b: dict, point_id: str, exclude_sid: str) -> List[str]:
    return [s for s in (b.get("allotments", {}).get(point_id, []) or []) if s != exclude_sid]


@api_router.post("/bandobasts/{bid}/alert")
async def send_bandobast_alert(bid: str):
    """Mark every allotted staff with a pending alert for this bandobast."""
    b = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Bandobast not found")
    if b.get("status") != "deployed":
        raise HTTPException(status_code=400, detail="Bandobast must be deployed before sending alerts")
    allot = b.get("allotments") or {}
    all_sids = set()
    for sids in allot.values():
        for s in sids:
            all_sids.add(s)
    if not all_sids:
        return {"ok": True, "sent": 0, "skipped_no_mobile": 0}
    sent = 0
    skipped_no_mobile = 0
    skipped_unknown = 0
    now = datetime.now(timezone.utc).isoformat()
    for sid in all_sids:
        staff = await _resolve_staff_for_bandobast(b, sid)
        if not staff:
            skipped_unknown += 1
            continue
        mobile = _normalize_mobile(staff.get("mobile") or "")
        if not mobile or len(mobile) != 10:
            skipped_no_mobile += 1
            continue
        await db.alerts.update_one(
            {"bandobast_id": bid, "staff_id": sid, "mobile": mobile},
            {"$set": {
                "bandobast_id": bid, "staff_id": sid, "mobile": mobile,
                "bandobast_name": b.get("name", ""), "bandobast_date": b.get("date", ""),
                "sent_at": now, "seen": False,
            }},
            upsert=True,
        )
        sent += 1
    await db.bandobasts.update_one({"id": bid}, {"$set": {"last_alerted_at": now}})
    return {"ok": True, "sent": sent, "skipped_no_mobile": skipped_no_mobile, "skipped_unknown": skipped_unknown}


@api_router.get("/bandobasts/{bid}/alert-status")
async def alert_status(bid: str):
    b = await db.bandobasts.find_one({"id": bid}, {"_id": 0, "last_alerted_at": 1, "allotments": 1, "other_district_staff": 1})
    if not b:
        raise HTTPException(status_code=404, detail="Not found")
    cnt = await db.alerts.count_documents({"bandobast_id": bid})
    seen = await db.alerts.count_documents({"bandobast_id": bid, "seen": True})
    return {"last_alerted_at": b.get("last_alerted_at"), "total": cnt, "seen": seen}


# ----- Staff App (mobile-number-based) -----

class StaffAppLogin(BaseModel):
    mobile: str


class StaffAppProfileUpdate(BaseModel):
    name: Optional[str] = None
    rank: Optional[str] = None
    posting: Optional[str] = None
    gender: Optional[str] = None
    district: Optional[str] = None
    category: Optional[str] = None
    photo: Optional[str] = None  # base64 data url


@api_router.post("/staff-app/login")
async def staff_app_login(payload: StaffAppLogin):
    mob = _normalize_mobile(payload.mobile)
    if len(mob) != 10:
        raise HTTPException(status_code=400, detail="Enter a valid 10-digit mobile number")
    matches = []
    async for s in db.staff.find({}, {"_id": 0}):
        if _normalize_mobile(s.get("mobile") or "") == mob:
            matches.append(s)
    if not matches:
        # Search OD staff inside bandobasts as a fallback
        async for b in db.bandobasts.find({}, {"_id": 0, "other_district_staff": 1}):
            for s in (b.get("other_district_staff") or []):
                if _normalize_mobile(s.get("mobile") or "") == mob:
                    matches.append(s)
    if not matches:
        raise HTTPException(status_code=404, detail="No staff found with this mobile number. Please contact admin.")
    return {"ok": True, "mobile": mob, "staff": matches[0]}


@api_router.get("/staff-app/me")
async def staff_app_me(mobile: str):
    mob = _normalize_mobile(mobile)
    s = None
    async for x in db.staff.find({}, {"_id": 0}):
        if _normalize_mobile(x.get("mobile") or "") == mob:
            s = x; break
    if not s:
        raise HTTPException(status_code=404, detail="Not found")
    return s


@api_router.patch("/staff-app/me")
async def staff_app_update_me(mobile: str, payload: StaffAppProfileUpdate):
    mob = _normalize_mobile(mobile)
    s = None
    async for x in db.staff.find({}, {"_id": 0}):
        if _normalize_mobile(x.get("mobile") or "") == mob:
            s = x; break
    if not s:
        raise HTTPException(status_code=404, detail="Not found")
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    update.pop("mobile", None)  # mobile is the auth key, never editable here
    if update:
        await db.staff.update_one({"id": s["id"]}, {"$set": update})
    return await db.staff.find_one({"id": s["id"]}, {"_id": 0})


@api_router.get("/staff-app/alerts")
async def staff_app_alerts(mobile: str):
    mob = _normalize_mobile(mobile)
    cur = db.alerts.find({"mobile": mob}, {"_id": 0}).sort("sent_at", -1)
    items = await cur.to_list(500)
    return items


@api_router.post("/staff-app/alerts/{bid}/seen")
async def staff_app_mark_alert_seen(bid: str, mobile: str):
    mob = _normalize_mobile(mobile)
    await db.alerts.update_many({"bandobast_id": bid, "mobile": mob}, {"$set": {"seen": True}})
    return {"ok": True}


@api_router.get("/staff-app/bandobast/{bid}")
async def staff_app_bandobast_detail(bid: str, mobile: str):
    """
    Returns everything the staff member needs in one call:
    bandobast meta, their assigned point with map link, equipment, suchana, and
    co-allotted staff at the same point.
    """
    mob = _normalize_mobile(mobile)
    b = await db.bandobasts.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Bandobast not found")
    # Find which staff record this mobile belongs to (home or OD)
    me = None
    async for x in db.staff.find({}, {"_id": 0}):
        if _normalize_mobile(x.get("mobile") or "") == mob:
            me = x; break
    if not me:
        for s in (b.get("other_district_staff") or []):
            if _normalize_mobile(s.get("mobile") or "") == mob:
                me = s; break
    if not me:
        raise HTTPException(status_code=404, detail="Staff not found")
    sid = me["id"]
    # Find the point I'm allotted to
    point = _point_for_staff(b, sid)
    if not point:
        return {
            "bandobast": {"id": b["id"], "name": b.get("name"), "date": b.get("date"), "spot": b.get("spot"), "in_charge": b.get("in_charge"), "ps_name": b.get("ps_name")},
            "me": me,
            "point": None,
            "equipment_for_me": None,
            "co_staff": [],
            "map_url": None,
        }
    eq = ((b.get("equipment_assignments") or {}).get(point["id"]) or {}).get(sid)
    co_ids = _co_staff_ids(b, point["id"], sid)
    co_staff = []
    for cid in co_ids:
        cs = await _resolve_staff_for_bandobast(b, cid)
        if cs:
            cs_eq = ((b.get("equipment_assignments") or {}).get(point["id"]) or {}).get(cid)
            co_staff.append({**cs, "equipment": cs_eq})
    map_url = None
    if point.get("latitude") is not None and point.get("longitude") is not None:
        map_url = f"https://www.google.com/maps?q={point['latitude']},{point['longitude']}"
    return {
        "bandobast": {"id": b["id"], "name": b.get("name"), "date": b.get("date"), "spot": b.get("spot"), "in_charge": b.get("in_charge"), "ps_name": b.get("ps_name")},
        "me": me,
        "point": point,
        "equipment_for_me": eq,
        "co_staff": co_staff,
        "map_url": map_url,
    }


# ========================= APP SETUP =========================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
